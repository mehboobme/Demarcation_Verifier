"""
Report Generator Module
Generates comprehensive validation reports in multiple formats.
"""

import os
import json
from datetime import datetime
from typing import List, Dict, Any
from dataclasses import asdict
import logging

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from validator import ValidationReport, ValidationResult, ValidationStatus

logging.basicConfig(level=logging.DEBUG)
logger = logging.getLogger(__name__)


# Color definitions
COLORS = {
    "header": "1F4E79",       # Dark blue
    "match": "C6EFCE",        # Light green
    "mismatch": "FFC7CE",     # Light red
    "warning": "FFEB9C",      # Light yellow
    "not_found": "D9D9D9",    # Light gray
    "critical": "FF0000",     # Red
    "high": "FFA500",         # Orange
    "medium": "FFD700",       # Gold
    "low": "90EE90",          # Light green
}


class ReportGenerator:
    """Generates validation reports in various formats"""
    
    def __init__(self, output_dir: str):
        self.output_dir = output_dir
        os.makedirs(output_dir, exist_ok=True)
        
    def generate_all_reports(self, reports: List[ValidationReport], filename_prefix: str = "validation"):
        """Generate Excel report only"""
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        
        # Generate Excel report only
        excel_path = os.path.join(self.output_dir, f"{filename_prefix}_{timestamp}.xlsx")
        self.generate_excel_report(reports, excel_path)
        
        return {
            "excel": excel_path
        }
    
    def generate_excel_report(self, reports: List[ValidationReport], output_path: str):
        """Generate comprehensive Excel report"""
        logger.info(f"Generating Excel report: {output_path}")
        
        wb = Workbook()
        
        # Create summary sheet
        self._create_summary_sheet(wb, reports)
        
        # Create page-by-page validation sheet
        self._create_page_by_page_sheet(wb, reports)
        
        # Create detailed results sheet
        self._create_details_sheet(wb, reports)
        
        # Create failures sheet
        self._create_failures_sheet(wb, reports)
        
        # Create comparison sheet
        self._create_comparison_sheet(wb, reports)
        
        # Remove default sheet if exists
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]
        
        wb.save(output_path)
        logger.info(f"Excel report saved: {output_path}")
    
    def _create_summary_sheet(self, wb: Workbook, reports: List[ValidationReport]):
        """Create summary sheet"""
        ws = wb.create_sheet("Summary", 0)
        
        # Styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color=COLORS["header"], end_color=COLORS["header"], fill_type="solid")
        
        # Title
        ws["A1"] = "PDF Validation Report"
        ws["A1"].font = Font(bold=True, size=16)
        ws.merge_cells("A1:F1")
        
        ws["A2"] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws["A3"] = f"Total PDFs Validated: {len(reports)}"
        
        # Statistics
        total_checks = sum(r.total_checks for r in reports)
        total_matches = sum(r.matches for r in reports)
        total_mismatches = sum(r.mismatches for r in reports)
        total_warnings = sum(r.warnings for r in reports)
        
        ws["A5"] = "Overall Statistics"
        ws["A5"].font = Font(bold=True, size=12)
        ws["A6"] = f"Total Checks: {total_checks}"
        ws["A7"] = f"Total Matches: {total_matches} ({100*total_matches/total_checks:.1f}%)" if total_checks else "Total Matches: 0"
        ws["A8"] = f"Total Mismatches: {total_mismatches}"
        ws["A9"] = f"Total Warnings: {total_warnings}"
        
        # Per-PDF summary table
        headers = ["PDF File", "Unit Code", "Plot #", "Status", "Matches", "Mismatches", "Warnings", "Critical", "High", "Medium"]
        row = 11
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        
        for report in reports:
            row += 1
            ws.cell(row=row, column=1, value=report.pdf_filename)
            ws.cell(row=row, column=2, value=report.unit_code)
            ws.cell(row=row, column=3, value=report.plot_number)
            
            status_cell = ws.cell(row=row, column=4, value=report.overall_status)
            if "CRITICAL" in report.overall_status:
                status_cell.fill = PatternFill(start_color=COLORS["critical"], end_color=COLORS["critical"], fill_type="solid")
            elif "HIGH" in report.overall_status:
                status_cell.fill = PatternFill(start_color=COLORS["high"], end_color=COLORS["high"], fill_type="solid")
            elif "PASSED" in report.overall_status:
                status_cell.fill = PatternFill(start_color=COLORS["match"], end_color=COLORS["match"], fill_type="solid")
            
            ws.cell(row=row, column=5, value=report.matches)
            ws.cell(row=row, column=6, value=report.mismatches)
            ws.cell(row=row, column=7, value=report.warnings)
            ws.cell(row=row, column=8, value=len(report.critical_failures))
            ws.cell(row=row, column=9, value=len(report.high_failures))
            ws.cell(row=row, column=10, value=len(report.medium_failures))
        
        # Adjust column widths
        for col in range(1, 11):
            ws.column_dimensions[get_column_letter(col)].width = 15
    
    def _create_page_by_page_sheet(self, wb: Workbook, reports: List[ValidationReport]):
        """Create page-by-page validation sheet organized by PDF page"""
        ws = wb.create_sheet("Page-by-Page Validation")
        
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color=COLORS["header"], end_color=COLORS["header"], fill_type="solid")
        section_font = Font(bold=True, size=11)
        section_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        
        # Define page groups for organization
        page_groups = {
            "Page 1": ["page1_unit_code"],
            "Page 2 Section 1": ["page2_sec1_nbhd_name", "page2_sec1_unit_type"],
            "Page 2 Section 2": ["page2_sec2_nbhd_name", "page2_sec2_unit_type", "page2_sec2_amana_plot_no"],
            "Page 3 Header": ["page3_header_unit_type"],
            "Page 3 (Dimensions)": ["page3_land_width", "page3_land_length", "page3_street_width"],
            "Page 3 Footer (Areas)": ["page3_footer_land_area", "page3_footer_gfa", "page3_footer_first_floor_area", 
                                       "page3_footer_ground_floor_area", "page3_footer_ground_floor_annex", 
                                       "page3_footer_second_floor_area"],
            "Page 4 Header": ["page4_header_unit_type", "page4_header_area"],
            "Page 5 Header": ["page5_header_unit_type", "page5_header_area"],
            "Page 6 Header": ["page6_header_unit_type", "page6_header_area"],
            "Page 7 Header": ["page7_header_unit_type"],
            "Page 8 Section 1 (Front)": ["page8_sec1_unit_type"],
            "Page 8 Section 2 (Back)": ["page8_sec2_unit_type"],
            "Page 9 Header (Side)": ["page9_header_unit_type"],
            "Typology Verification": ["typology_page4_ground", "typology_page5_first", "typology_page6_top", 
                                      "typology_page7_terrace", "typology_page8_front", "typology_page9_back"],
            "Consolidated": ["unit_code", "plot_number", "unit_type", "land_area", "ground_floor_area", 
                            "first_floor_area", "upper_floor_area", "total_floor_area", "ground_floor_annex",
                            "land_length", "land_width", "street_width", "facade_type",
                            "red_highlight_neighborhood", "red_highlight_plot", "corner_unit", "connected_streets"]
        }
        
        # Headers
        headers = ["PDF File", "Plot #", "Page/Section", "Field", "PDF Field (Arabic)", "PDF Value", 
                   "Excel Column", "Ground Truth", "Status", "Match?"]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
        
        # Field to Arabic/Excel mapping
        field_mapping = {
            "page1_unit_code": ("رقم الوحدة السكنية", "Unit_Code"),
            "page2_sec1_nbhd_name": ("المنطقة", "NBHD_Name"),
            "page2_sec1_unit_type": ("الخريطة المرجعية", "Unit_Type"),
            "page2_sec2_nbhd_name": ("المنطقة", "NBHD_Name"),
            "page2_sec2_unit_type": ("الخريطة المرجعية", "Unit_Type"),
            "page2_sec2_amana_plot_no": ("القطعة رقم", "AMANA_Plot_No"),
            "page3_header_unit_type": ("الموقع العام", "Unit_Type"),
            "page3_land_width": ("plot width", "Land_Width"),
            "page3_land_length": ("plot length", "Land_Length"),
            "page3_street_width": ("عرض الشارع", "Street_Width"),
            "page3_footer_land_area": ("مساحة الأرض", "Land_Area"),
            "page3_footer_gfa": ("إجمالي مساحة الطوابق", "GFA"),
            "page3_footer_first_floor_area": ("مساحة الطابق الأول", "First_Floor_Area"),
            "page3_footer_ground_floor_area": ("مساحة الطابق الأرضي", "Ground_Floor_Area"),
            "page3_footer_ground_floor_annex": ("مساحة الملحق الأرضي", "Ground_Floor_Annex_Driver"),
            "page3_footer_second_floor_area": ("مساحة الطابق العلوي", "Second_Floor_Area"),
            "page4_header_unit_type": ("النموذج,الطابق الأرضي مع الملحق", "Unit_Type"),
            "page4_header_area": ("المساحة,الطابق الأرضي مع الملحق", "Ground_Floor_Area+Ground_Floor_Annex"),
            "page5_header_unit_type": ("النموذج,الطابق الأول", "Unit_Type"),
            "page5_header_area": ("المساحة,الطابق الأول", "First_Floor_Area"),
            "page6_header_unit_type": ("النموذج,الملحق العلوي", "Unit_Type"),
            "page6_header_area": ("المساحة,الملحق العلوي", "Second_Floor_Area"),
            "page7_header_unit_type": ("النموذج,السطح", "Unit_Type"),
            "page8_sec1_unit_type": ("الواجهة الأمامية,النموذج", "Unit_Type"),
            "page8_sec2_unit_type": ("النموذج,الواجهة الخلفية", "Unit_Type"),
            "page9_header_unit_type": ("النموذج,الواجهة الجانبية", "Unit_Type"),
            "unit_code": ("Unit Code", "Unit_Code"),
            "plot_number": ("Plot Number", "AMANA_Plot_No"),
            "unit_type": ("Unit Type", "Unit_Type"),
            "land_area": ("Land Area", "Land_Area"),
            "ground_floor_area": ("Ground Floor Area", "Ground_Floor_Area"),
            "first_floor_area": ("First Floor Area", "First_Floor_Area"),
            "upper_floor_area": ("Upper Floor Area", "Second_Floor_Area"),
            "total_floor_area": ("Total Floor Area (GFA)", "GFA"),
            "ground_floor_annex": ("Ground Floor Annex", "Ground_Floor_Annex_Driver"),
            "land_length": ("Land Length", "Land_Length"),
            "land_width": ("Land Width", "Land_Width"),
            "street_width": ("Street Width", "Street_Width"),
            "facade_type": ("Facade Type", "Façade_Type"),
            "red_highlight_neighborhood": ("Red Highlight (Neighborhood)", "Visual Check"),
            "red_highlight_plot": ("Red Highlight (Plot)", "Visual Check"),
            "corner_unit": ("Corner Unit", "Corner_Unit"),
            "connected_streets": ("Number of Streets", "Number_Of_Streets"),
            "typology_page4_ground": ("Ground Floor Plan Match", "Unit_Type"),
            "typology_page5_first": ("First Floor Plan Match", "Unit_Type"),
            "typology_page6_top": ("Top Floor Plan Match", "Unit_Type"),
            "typology_page7_terrace": ("Terrace Plan Match", "Unit_Type"),
            "typology_page8_front": ("Front Facade Match", "Unit_Type"),
            "typology_page9_back": ("Back Facade Match", "Unit_Type"),
        }
        
        row = 2
        for report in reports:
            # Create a lookup dict for this report's results
            result_lookup = {r.field_name: r for r in report.all_results}
            
            for page_name, field_list in page_groups.items():
                # Page section header
                ws.cell(row=row, column=1, value=report.pdf_filename)
                ws.cell(row=row, column=2, value=report.plot_number)
                section_cell = ws.cell(row=row, column=3, value=page_name)
                section_cell.font = section_font
                section_cell.fill = section_fill
                ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=10)
                row += 1
                
                for field_name in field_list:
                    result = result_lookup.get(field_name)
                    if result:
                        ws.cell(row=row, column=1, value=report.pdf_filename)
                        ws.cell(row=row, column=2, value=report.plot_number)
                        ws.cell(row=row, column=3, value=page_name)
                        ws.cell(row=row, column=4, value=field_name)
                        
                        # Arabic name and Excel column
                        arabic_name, excel_col = field_mapping.get(field_name, (field_name, field_name))
                        ws.cell(row=row, column=5, value=arabic_name)
                        ws.cell(row=row, column=6, value=str(result.pdf_value))
                        ws.cell(row=row, column=7, value=excel_col)
                        ws.cell(row=row, column=8, value=str(result.ground_truth_value))
                        ws.cell(row=row, column=9, value=result.status.value)
                        
                        # Match indicator with color
                        match_cell = ws.cell(row=row, column=10, value="✓" if result.status == ValidationStatus.MATCH else "✗")
                        if result.status == ValidationStatus.MATCH:
                            match_cell.fill = PatternFill(start_color=COLORS["match"], end_color=COLORS["match"], fill_type="solid")
                        elif result.status == ValidationStatus.MISMATCH:
                            match_cell.fill = PatternFill(start_color=COLORS["mismatch"], end_color=COLORS["mismatch"], fill_type="solid")
                        elif result.status == ValidationStatus.NOT_FOUND:
                            match_cell.fill = PatternFill(start_color=COLORS["not_found"], end_color=COLORS["not_found"], fill_type="solid")
                        
                        row += 1
                
            row += 1  # Add spacing between PDFs
        
        # Adjust column widths
        widths = [30, 8, 22, 28, 30, 15, 30, 15, 12, 8]
        for col, width in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = width

    def _create_details_sheet(self, wb: Workbook, reports: List[ValidationReport]):
        """Create detailed results sheet"""
        ws = wb.create_sheet("Detailed Results")
        
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color=COLORS["header"], end_color=COLORS["header"], fill_type="solid")
        
        headers = ["PDF File", "Unit Code", "Plot #", "Field", "PDF Value", "Ground Truth", "Status", "Severity", "Message"]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
        
        row = 2
        for report in reports:
            for result in report.all_results:
                ws.cell(row=row, column=1, value=report.pdf_filename)
                ws.cell(row=row, column=2, value=report.unit_code)
                ws.cell(row=row, column=3, value=report.plot_number)
                ws.cell(row=row, column=4, value=result.field_name)
                ws.cell(row=row, column=5, value=str(result.pdf_value))
                ws.cell(row=row, column=6, value=str(result.ground_truth_value))
                
                status_cell = ws.cell(row=row, column=7, value=result.status.value)
                if result.status == ValidationStatus.MATCH:
                    status_cell.fill = PatternFill(start_color=COLORS["match"], end_color=COLORS["match"], fill_type="solid")
                elif result.status == ValidationStatus.MISMATCH:
                    status_cell.fill = PatternFill(start_color=COLORS["mismatch"], end_color=COLORS["mismatch"], fill_type="solid")
                elif result.status == ValidationStatus.WARNING:
                    status_cell.fill = PatternFill(start_color=COLORS["warning"], end_color=COLORS["warning"], fill_type="solid")
                elif result.status == ValidationStatus.NOT_FOUND:
                    status_cell.fill = PatternFill(start_color=COLORS["not_found"], end_color=COLORS["not_found"], fill_type="solid")
                
                severity_names = {1: "CRITICAL", 2: "HIGH", 3: "MEDIUM", 4: "LOW"}
                ws.cell(row=row, column=8, value=severity_names.get(result.severity, "UNKNOWN"))
                ws.cell(row=row, column=9, value=result.message)
                
                row += 1
        
        # Adjust column widths
        widths = [30, 25, 10, 25, 20, 20, 12, 12, 50]
        for col, width in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = width
    
    def _create_failures_sheet(self, wb: Workbook, reports: List[ValidationReport]):
        """Create failures-only sheet"""
        ws = wb.create_sheet("Failures Only")
        
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color=COLORS["header"], end_color=COLORS["header"], fill_type="solid")
        
        headers = ["PDF File", "Unit Code", "Plot #", "Field", "PDF Value", "Ground Truth", "Severity", "Message"]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
        
        row = 2
        for report in reports:
            all_failures = (
                report.critical_failures + 
                report.high_failures + 
                report.medium_failures + 
                report.low_failures
            )
            
            for result in all_failures:
                ws.cell(row=row, column=1, value=report.pdf_filename)
                ws.cell(row=row, column=2, value=report.unit_code)
                ws.cell(row=row, column=3, value=report.plot_number)
                ws.cell(row=row, column=4, value=result.field_name)
                ws.cell(row=row, column=5, value=str(result.pdf_value))
                ws.cell(row=row, column=6, value=str(result.ground_truth_value))
                
                severity_names = {1: "CRITICAL", 2: "HIGH", 3: "MEDIUM", 4: "LOW"}
                severity_cell = ws.cell(row=row, column=7, value=severity_names.get(result.severity, "UNKNOWN"))
                
                if result.severity == 1:
                    severity_cell.fill = PatternFill(start_color=COLORS["critical"], end_color=COLORS["critical"], fill_type="solid")
                elif result.severity == 2:
                    severity_cell.fill = PatternFill(start_color=COLORS["high"], end_color=COLORS["high"], fill_type="solid")
                elif result.severity == 3:
                    severity_cell.fill = PatternFill(start_color=COLORS["medium"], end_color=COLORS["medium"], fill_type="solid")
                
                ws.cell(row=row, column=8, value=result.message)
                row += 1
        
        # Adjust column widths
        widths = [30, 25, 10, 25, 20, 20, 12, 50]
        for col, width in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = width
    
    def _create_comparison_sheet(self, wb: Workbook, reports: List[ValidationReport]):
        """Create side-by-side comparison sheet"""
        ws = wb.create_sheet("PDF vs Ground Truth")
        
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color=COLORS["header"], end_color=COLORS["header"], fill_type="solid")
        
        headers = ["Field", "PDF Value", "Ground Truth Value", "Match?", "Difference"]
        
        row = 1
        for report in reports:
            # Section header
            ws.cell(row=row, column=1, value=f"PDF: {report.pdf_filename}")
            ws.cell(row=row, column=1).font = Font(bold=True, size=12)
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
            row += 1
            
            # Column headers
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
            row += 1
            
            # Results
            for result in report.all_results:
                ws.cell(row=row, column=1, value=result.field_name)
                ws.cell(row=row, column=2, value=str(result.pdf_value))
                ws.cell(row=row, column=3, value=str(result.ground_truth_value))
                
                match_cell = ws.cell(row=row, column=4, value="✓" if result.status == ValidationStatus.MATCH else "✗")
                if result.status == ValidationStatus.MATCH:
                    match_cell.fill = PatternFill(start_color=COLORS["match"], end_color=COLORS["match"], fill_type="solid")
                else:
                    match_cell.fill = PatternFill(start_color=COLORS["mismatch"], end_color=COLORS["mismatch"], fill_type="solid")
                
                # Calculate difference for numeric values
                try:
                    pdf_val = float(result.pdf_value)
                    gt_val = float(result.ground_truth_value)
                    diff = pdf_val - gt_val
                    ws.cell(row=row, column=5, value=f"{diff:+.4f}")
                except:
                    ws.cell(row=row, column=5, value="-")
                
                row += 1
            
            row += 2  # Add spacing between PDFs
        
        # Adjust column widths
        widths = [25, 20, 20, 10, 15]
        for col, width in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = width
    
    def generate_json_report(self, reports: List[ValidationReport], output_path: str):
        """Generate JSON report"""
        logger.info(f"Generating JSON report: {output_path}")
        
        def serialize_result(result: ValidationResult) -> Dict:
            return {
                "field_name": result.field_name,
                "pdf_value": str(result.pdf_value),
                "ground_truth_value": str(result.ground_truth_value),
                "status": result.status.value,
                "severity": result.severity,
                "message": result.message
            }
        
        data = {
            "generated_at": datetime.now().isoformat(),
            "total_pdfs": len(reports),
            "reports": []
        }
        
        for report in reports:
            report_data = {
                "pdf_filename": report.pdf_filename,
                "unit_code": report.unit_code,
                "plot_number": report.plot_number,
                "overall_status": report.overall_status,
                "statistics": {
                    "total_checks": report.total_checks,
                    "matches": report.matches,
                    "mismatches": report.mismatches,
                    "warnings": report.warnings,
                    "not_found": report.not_found
                },
                "critical_failures": [serialize_result(r) for r in report.critical_failures],
                "high_failures": [serialize_result(r) for r in report.high_failures],
                "medium_failures": [serialize_result(r) for r in report.medium_failures],
                "all_results": [serialize_result(r) for r in report.all_results]
            }
            data["reports"].append(report_data)
        
        with open(output_path, "w", encoding="utf-8") as f:
            json.dump(data, f, indent=2, ensure_ascii=False)
        
        logger.info(f"JSON report saved: {output_path}")
    
    def generate_text_report(self, reports: List[ValidationReport], output_path: str):
        """Generate text summary report"""
        logger.info(f"Generating text report: {output_path}")
        
        lines = [
            "=" * 80,
            "PDF VALIDATION REPORT",
            "=" * 80,
            f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
            f"Total PDFs Validated: {len(reports)}",
            "",
        ]
        
        # Overall statistics
        total_checks = sum(r.total_checks for r in reports)
        total_matches = sum(r.matches for r in reports)
        total_mismatches = sum(r.mismatches for r in reports)
        
        lines.extend([
            "OVERALL STATISTICS",
            "-" * 40,
            f"Total Checks: {total_checks}",
            f"Total Matches: {total_matches}",
            f"Total Mismatches: {total_mismatches}",
            f"Match Rate: {100*total_matches/total_checks:.1f}%" if total_checks else "N/A",
            "",
        ])
        
        # Per-PDF details
        for report in reports:
            lines.append(report.summary)
            lines.append("")
        
        with open(output_path, "w", encoding="utf-8") as f:
            f.write("\n".join(lines))
        
        logger.info(f"Text report saved: {output_path}")


if __name__ == "__main__":
    # Test with sample data
    from validator import ValidationReport, ValidationResult, ValidationStatus
    
    # Create sample report
    sample_report = ValidationReport(
        pdf_filename="test.pdf",
        unit_code="DM1-D02-2A-25-453",
        plot_number=453,
        total_checks=10,
        matches=8,
        mismatches=2
    )
    
    generator = ReportGenerator("/home/claude/pdf_validator/output")
    paths = generator.generate_all_reports([sample_report], "test")
    print(f"Reports generated: {paths}")
